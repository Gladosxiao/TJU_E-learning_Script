# -*- coding: utf-8 -*-
"""
Created on Fri Apr 10 12:30:50 2020

@author: GlaDo
"""

import  os

from openpyxl import load_workbook,Workbook

try:
    ListD=[] 
    N = 0
    
    for root, dirs, files in os.walk(os.getcwd()): 
        for file in files: 
            if os.path.splitext(file)[1] == '.xlsx': 
                ListD.append(os.path.join(root, file)) ##读取当前目录下所有文件名称
    
    wbName = load_workbook(filename = ListD[0])##打开输入的文件 读学生学号
    wbName_sheet = wbName['Sheet1']
    ListName=[]
    ListNo=[]  
    for i in range(len(wbName_sheet["A"])):#学号和姓名读入程序
        ListNo.append(wbName_sheet["A"][i].value)
        ListName.append(wbName_sheet["B"][i].value)
    
    wbName.close()#关闭文件    
    l = len(ListNo)
    ListAttend=[0]*l
    ListLate=[0]*l
    ListBackup=[""]*l
    del ListD[0]#删除学号与姓名文件的路径     
    
    for i in ListD:#开始对每个文件进行操作
               
        wbWorking = load_workbook(filename = i)##打开本次操作的文件     
        wbWorking_sheet = wbWorking['Sheet1']#仅支持新版的出勤统计表格 老版本请自行格式化或是人工处理
    
        if wbWorking_sheet['D1'].value != "时长":
            N += 1
            continue#如果是老格式版本的表格 自动调过
        ListNameonetime= list(ListName)
        for j in range(1,wbWorking_sheet.max_row,1):#按照行读入一行信息 
            Flag = 1#查找学生
            Namestring = wbWorking_sheet["B"][j].value#读取学生姓名
            #由于学生姓名填写的不确定性，以每个姓名的字符都在学生自行填写的与会者中出现过 的方式做匹配
            for q in ListNameonetime:
                NameFlag = 0#查找学生
                for p in q:
                    if p in Namestring:
                        NameFlag += 1
                if NameFlag==len(q):
                    index = ListName.index(q)#记录index为写入做准备
                    ListNameonetime.remove(q)#已经查到了的学生从列表删除
                    Flag = 0
                    break#找到就结束循环
            if Flag:
                continue
                
            
            if int(wbWorking_sheet["D"][j].value)>91:#时长超过91min的视为出席
                ListAttend[index] = ListAttend[index]+1
            else:
                ListLate[index] = ListLate[index]+1
                ListBackup[index] =ListBackup[index] +" "+ wbWorking_sheet["E"][j].value            
                    
        wbWorking.close()#关闭文件    
    
    wbout = Workbook()#文件写入
    wbbout_sheet = wbout.active
    wbbout_sheet["A1"].value = "学号"
    wbbout_sheet["B1"].value = "姓名"
    wbbout_sheet["C1"].value = "出席"
    wbbout_sheet["D1"].value = "迟到"
    wbbout_sheet["E1"].value = "缺勤"
    wbbout_sheet["F1"].value = "备注"
    
    for i in range(len(ListName)):
        wbbout_sheet["A"+str(i+2)].value = ListNo[i]
        wbbout_sheet["B"+str(i+2)].value = ListName[i]
        wbbout_sheet["C"+str(i+2)].value = ListAttend[i]
        wbbout_sheet["D"+str(i+2)].value = ListLate[i]
        wbbout_sheet["E"+str(i+2)].value = len(ListD)-ListAttend[i]-ListLate[i]-N
        wbbout_sheet["F"+str(i+2)].value = ListBackup[i]
        
    wbout.save(file[16:-5]+"_"+"出勤统计.xlsx")    

         
    
                
except:
    print("新版本表格统不成功")
else:
    print("新版本表格统计成功，老版本表格"+str(N)+"个未进行统计"+"总共统计上课"+str(len(ListD)-N)+"次")
    