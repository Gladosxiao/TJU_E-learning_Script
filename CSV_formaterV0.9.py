# -*- coding: utf-8 -*-
"""
Created on Fri Apr 10 12:30:50 2020

@author: GlaDo
"""

import  os
#import openpyxl
import csv

import openpyxl

try:
    ListD=[]
    

    for root, dirs, files in os.walk(os.getcwd()): 
        for file in files: 
            if os.path.splitext(file)[1] == '.csv': 
                ListD.append(os.path.join(root, file)) 
    for i in ListD:
        ListX=[]
        wb = openpyxl.Workbook()
        work_sheet = wb.create_sheet(title=u"Sheet1")
        # 打开csv文件
        csvfile = open(i)
 
        # 获取csv.reader
        lines = csv.reader(csvfile)
        row = 1
        for line in lines:
            # print(line)
            lin = 1
            for i in line:
                work_sheet.cell(row=row, column=lin).value = i
                lin += 1
            row += 1
        csvfile.close()
     
        Name = work_sheet["B2"].value
        work_sheet.delete_rows(1,3) # 删除第1-3行   
        work_sheet.delete_cols(2,9) 
        work_sheet.delete_cols(4,37) 
        for j in range(1,(work_sheet.max_row-1)):
            if work_sheet["A"+str(j)].value == '':
                ListX.append(j)
            else:
                work_sheet["C"+str(j)].value = work_sheet["C"+str(j)].value[0:8]
        for p in range(len(ListX)-1):
            work_sheet.delete_rows(ListX[p])  
            ListX[p+1] = ListX[p+1]-p-1
            
        work_sheet.delete_rows(ListX[p+1]) 
        work_sheet.delete_rows(ListX[p+1]+1) 
        work_sheet.insert_cols(1)
        work_sheet["A1"] = "上课时间"
        for q in range(2,(work_sheet.max_row+1)):
            work_sheet["A"+str(q)].value = Name[0:10]
        
        Name = Name.replace(' ','_')
        Name = Name.replace('-','_')
        Name = Name.replace(':','')
        wb.remove(wb["Sheet"])
        wb.save(Name+'.xlsx')
            

         
    
                
except:
    print("格式修改不成功")
else:
    print("格式修改成功")
    