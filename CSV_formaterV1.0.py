# -*- coding: utf-8 -*-
"""
Created on Fri Apr 10 12:30:50 2020

@author: GlaDo
"""

import  os
#import openpyxl
import csv

import openpyxl

#try:
ListD=[]


for root, dirs, files in os.walk(os.getcwd()): 
    for file in files: 
        if os.path.splitext(file)[1] == '.csv': 
            ListD.append(os.path.join(root, file)) 
for i in ListD:
    ListX=[]
    wb = openpyxl.Workbook()
    work_sheet = wb.create_sheet(title=u"Sheet1")
    # 打开csv文件
    csvfile = open(i)
 
    # 获取csv.reader
    lines = csv.reader(csvfile)
    row = 1
    for line in lines:
        # print(line)
        lin = 1
        for i in line:
            work_sheet.cell(row=row, column=lin).value = i
            lin += 1
        row += 1
    csvfile.close()
 
    Name = work_sheet["B2"].value
    work_sheet.delete_rows(1,3) # 删除第1-3行   
    work_sheet.delete_cols(2,9) 
    work_sheet.delete_cols(4,37) 
    for j in range(1,(work_sheet.max_row-1)):
        if work_sheet["A"+str(j)].value == '':
            ListX.append(j)
        else:
            work_sheet["C"+str(j)].value = work_sheet["C"+str(j)].value[0:8]
    for p in range(len(ListX)-1):
        work_sheet.delete_rows(ListX[p])  
        ListX[p+1] = ListX[p+1]-p-1
        
    work_sheet.delete_rows(ListX[p+1]) 
    work_sheet.delete_rows(ListX[p+1]+1) 
    work_sheet.insert_cols(1)
    work_sheet.insert_cols(3,2) #插入两个空列
    work_sheet["A1"] = "上课时间"
    work_sheet["D1"] = "时长"
    for q in range(1,(work_sheet.max_row)): #计算时长

        FF = work_sheet["F"][q].value.find(":")
        EE = work_sheet["E"][q].value.find(":")
        hour= int(work_sheet["F"][q].value[:FF])-int(work_sheet["E"][q].value[:EE])
        minute = int(work_sheet["F"][q].value[FF+1:FF+3])-int(work_sheet["E"][q].value[EE+1:EE+3])#截取字符串获得时长
        work_sheet["D"][q].value = hour*60+minute
    for q in range(2,(work_sheet.max_row+1)):
        work_sheet["A"+str(q)].value = Name[0:10]
    
    for q in range(2,work_sheet.max_row-1): 
        for p in range(q+1,work_sheet.max_row): 
            if work_sheet["D"][p].value >0 and work_sheet["B"][p].value == work_sheet["B"][q].value:

                work_sheet["D"][q].value = work_sheet["D"][q].value+work_sheet["D"][p].value
                work_sheet["D"][p].value = -1#找出重复条目
    j = 2#
    while j<work_sheet.max_row:#删除重复条目
        if  work_sheet["D"][j].value == -1:
            work_sheet.delete_rows(j+1) 
            j -= 1
        j += 1
            
            
    
    Name = Name.replace(' ','_')
    Name = Name.replace('-','_')
    Name = Name.replace(':','')
    wb.remove(wb["Sheet"])
    wb.save(Name+'.xlsx')
                
#except:
 #   print("格式修改不成功")
#else:
 #   print("格式修改成功")
    